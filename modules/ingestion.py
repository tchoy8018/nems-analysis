from pathlib import Path
from datetime import datetime, timezone

import pandas as pd
from sqlalchemy import text


EXCEL_COLUMN_MAP = {
    "Source.Name": "source_file",
    "DATE": "date",
    "PERIOD": "period",
    "USEP ($/MWh)": "usep",
    "LCP ($/MWh)": "lcp",
    "DEMAND (MW)": "demand_mw",
    "SOLAR(MW)": "solar_mw",
    "TCL (MW)": "tcl_mw",
    "RUSEP ($/MWh)": "rusep",
    "MAP ($/MWh)": "map_price",
    "MAPT ($/MWh)": "mapt_price",
    "TPC Applied": "tpc_applied",
}

NUMERIC_COLS = ["usep", "lcp", "demand_mw", "solar_mw", "tcl_mw",
                "rusep", "map_price", "mapt_price"]

CHUNKSIZE = 5000


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Rename columns, coerce types, and drop rows missing date/period."""
    present = {k: v for k, v in EXCEL_COLUMN_MAP.items() if k in df.columns}
    df = df.rename(columns=present)

    required = {"date", "period"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Required columns missing after rename: {missing}")

    if pd.api.types.is_datetime64_any_dtype(df["date"]):
        df["date"] = df["date"].dt.date
    else:
        df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date

    df["period"] = pd.to_numeric(df["period"], errors="coerce")
    df = df.dropna(subset=["date", "period"])
    df["period"] = df["period"].astype(int)

    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        else:
            df[col] = float("nan")

    if "tpc_applied" not in df.columns:
        df["tpc_applied"] = None
    if "source_file" not in df.columns:
        df["source_file"] = None

    df["imported_at"] = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()

    keep = [
        "source_file", "date", "period", "usep", "lcp", "demand_mw",
        "solar_mw", "tcl_mw", "rusep", "map_price", "mapt_price",
        "tpc_applied", "imported_at",
    ]
    return df[[c for c in keep if c in df.columns]]


def _detect_dialect(engine) -> str:
    return engine.dialect.name  # "sqlite" or "postgresql"


def _insert_chunk(chunk: pd.DataFrame, engine) -> tuple[int, int]:
    """Insert a chunk, returning (rows_inserted, rows_skipped)."""
    dialect = _detect_dialect(engine)
    records = chunk.to_dict(orient="records")

    inserted = 0
    skipped = 0

    with engine.begin() as conn:
        for row in records:
            if dialect == "sqlite":
                sql = text("""
                    INSERT OR IGNORE INTO nems_prices
                    (source_file, date, period, usep, lcp, demand_mw, solar_mw,
                     tcl_mw, rusep, map_price, mapt_price, tpc_applied, imported_at)
                    VALUES
                    (:source_file, :date, :period, :usep, :lcp, :demand_mw, :solar_mw,
                     :tcl_mw, :rusep, :map_price, :mapt_price, :tpc_applied, :imported_at)
                """)
            else:
                sql = text("""
                    INSERT INTO nems_prices
                    (source_file, date, period, usep, lcp, demand_mw, solar_mw,
                     tcl_mw, rusep, map_price, mapt_price, tpc_applied, imported_at)
                    VALUES
                    (:source_file, :date, :period, :usep, :lcp, :demand_mw, :solar_mw,
                     :tcl_mw, :rusep, :map_price, :mapt_price, :tpc_applied, :imported_at)
                    ON CONFLICT (date, period) DO NOTHING
                """)

            result = conn.execute(sql, row)
            if result.rowcount == 1:
                inserted += 1
            else:
                skipped += 1

    return inserted, skipped


def import_excel_to_db(excel_path: Path, engine) -> dict:
    """
    Import the NEMS master Excel file into the database.

    Returns a summary dict with keys:
        rows_imported, rows_skipped, date_range (min, max), errors
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        return {
            "rows_imported": 0,
            "rows_skipped": 0,
            "date_range": None,
            "errors": [f"File not found: {excel_path}"],
        }

    errors = []
    total_imported = 0
    total_skipped = 0
    all_dates = []

    print(f"Reading {excel_path.name} ...")
    try:
        raw = pd.read_excel(
            excel_path,
            sheet_name="datasource Market USEP",
            engine="openpyxl",
        )
    except Exception as e:
        return {
            "rows_imported": 0,
            "rows_skipped": 0,
            "date_range": None,
            "errors": [f"Failed to read Excel: {e}"],
        }

    print(f"  Total rows read: {len(raw):,}")

    try:
        df = _clean_dataframe(raw)
    except Exception as e:
        return {
            "rows_imported": 0,
            "rows_skipped": 0,
            "date_range": None,
            "errors": [f"Data cleaning failed: {e}"],
        }

    print(f"  Rows after cleaning: {len(df):,}")

    for chunk_start in range(0, len(df), CHUNKSIZE):
        chunk = df.iloc[chunk_start : chunk_start + CHUNKSIZE]
        chunk_num = chunk_start // CHUNKSIZE + 1
        try:
            ins, skip = _insert_chunk(chunk, engine)
            total_imported += ins
            total_skipped += skip
            all_dates.extend(chunk["date"].tolist())
            print(f"  Chunk {chunk_num}: +{ins} inserted, {skip} skipped")
        except Exception as e:
            errors.append(f"Chunk {chunk_num}: {e}")
            print(f"  Chunk {chunk_num} ERROR: {e}")

    date_range = None
    if all_dates:
        valid_dates = [d for d in all_dates if d is not None]
        if valid_dates:
            date_range = {"min": min(valid_dates), "max": max(valid_dates)}

    return {
        "rows_imported": total_imported,
        "rows_skipped": total_skipped,
        "date_range": date_range,
        "errors": errors,
    }


def import_csv_to_db(csv_path: Path, engine) -> dict:
    """
    Append a single EMC monthly CSV to the database.
    Deduplicates by (date, period).
    """
    csv_path = Path(csv_path)
    if not csv_path.exists():
        return {
            "rows_imported": 0,
            "rows_skipped": 0,
            "date_range": None,
            "errors": [f"File not found: {csv_path}"],
        }

    errors = []
    try:
        raw = pd.read_csv(csv_path)
        df = _clean_dataframe(raw)
        ins, skip = _insert_chunk(df, engine)
        valid_dates = [d for d in df["date"].tolist() if d is not None]
        date_range = {"min": min(valid_dates), "max": max(valid_dates)} if valid_dates else None
        return {
            "rows_imported": ins,
            "rows_skipped": skip,
            "date_range": date_range,
            "errors": errors,
        }
    except Exception as e:
        return {
            "rows_imported": 0,
            "rows_skipped": 0,
            "date_range": None,
            "errors": [str(e)],
        }


def detect_file_type(df: pd.DataFrame) -> str:
    """
    Infer file type from column names.
    Returns: 'nems_prices' | 'gas_prices' | 'analyst_forecast' | 'unknown'
    """
    cols = {c.lower().strip() for c in df.columns}
    if any(c in cols for c in ("usep ($/mwh)", "usep", "lcp ($/mwh)")):
        return "nems_prices"
    if any(c in cols for c in ("jkm", "jkm_usd_mmbtu", "piped_gas", "piped_gas_sgd_mmbtu")):
        return "gas_prices"
    if any(c in cols for c in ("forecast_price", "price", "usep_forecast", "annual_avg", "monthly_avg")):
        return "analyst_forecast"
    return "unknown"


def _parse_spglobal_gas_sheet(ws) -> dict:
    """
    Parse one sheet from the S&P Global Singapore Gas Trade Data workbook.
    Returns: {(year, month): {vol_mt, val_usd, price_usd_mmbtu}}
    Row layout (1-indexed): 5=years, 7-18=vol, 20-31=val, 33-44=price
    """
    rows = list(ws.iter_rows(min_row=1, max_row=44, values_only=True))

    # Year headers: row 5 (index 4), starting col C (index 2)
    year_row = rows[4]
    year_cols: list[tuple[int, int]] = []   # (col_index, year)
    for ci, val in enumerate(year_row):
        if ci < 2 or val is None:
            continue
        try:
            yr = int(str(val).strip())
            year_cols.append((ci, yr))
        except ValueError:
            pass

    # Block offsets (0-indexed first data row): vol=6, val=19, price=32
    MONTHS = ["January","February","March","April","May","June",
              "July","August","September","October","November","December"]
    MONTH_NUM = {m: i+1 for i, m in enumerate(MONTHS)}

    data: dict[tuple[int,int], dict] = {}

    for block_start, block_key in [(6, "vol"), (19, "val"), (32, "price")]:
        for month_offset in range(12):
            row = rows[block_start + month_offset]
            month_label = row[1]
            if month_label not in MONTH_NUM:
                continue
            month = MONTH_NUM[month_label]
            for ci, year in year_cols:
                val = row[ci] if ci < len(row) else None
                if val is None:
                    continue
                try:
                    val = float(val)
                except (TypeError, ValueError):
                    continue
                key = (year, month)
                if key not in data:
                    data[key] = {}
                data[key][block_key] = val

    return data


def ingest_gas_prices(excel_path, engine, fx_usd_sgd: float = 1.35) -> dict:
    """
    Parse Singapore Gas Trade Data workbook (S&P Global / SG Customs).
    Reads Malaysia piped, Indonesia piped, and LNG import sheets.
    Computes weighted average price and implied CCGT cost floor.

    Conversion factors: LNG 1 MT = 52 MMBtu; piped gas 1 MT = 50 MMBtu
    Heat rate: 7.5 MMBtu/MWh (standard Singapore CCGT)
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        return {"rows_ingested": 0, "errors": [f"File not found: {excel_path}"]}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        return {"rows_ingested": 0, "errors": [str(e)]}

    SHEET_MAP = {
        "malaysia":  "Piped gas import from Malaysia",
        "indonesia": "Piped gas import from Indonesia",
        "lng":       "LNG imports",
    }
    MMBTU_PER_MT = {"malaysia": 50.0, "indonesia": 50.0, "lng": 52.0}

    parsed: dict[str, dict] = {}
    for src, sheet_name in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            parsed[src] = {}
            continue
        parsed[src] = _parse_spglobal_gas_sheet(wb[sheet_name])

    # Merge into monthly records
    all_keys: set[tuple[int,int]] = set()
    for src_data in parsed.values():
        all_keys.update(src_data.keys())

    now_str = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
    dialect = _detect_dialect(engine)
    inserted = 0
    skipped = 0
    errors: list[str] = []
    all_dates = []

    with engine.begin() as conn:
        for year, month in sorted(all_keys):
            try:
                price_date = f"{year:04d}-{month:02d}-01"

                my = parsed["malaysia"].get((year, month), {})
                id_ = parsed["indonesia"].get((year, month), {})
                lg  = parsed["lng"].get((year, month), {})

                # MMBtu volumes for weighted avg
                my_mmbtu  = (my.get("vol", 0) or 0) * MMBTU_PER_MT["malaysia"]
                id_mmbtu  = (id_.get("vol", 0) or 0) * MMBTU_PER_MT["indonesia"]
                lng_mmbtu = (lg.get("vol", 0) or 0)  * MMBTU_PER_MT["lng"]
                total_mmbtu = my_mmbtu + id_mmbtu + lng_mmbtu
                total_vol_mt = (
                    (my.get("vol") or 0) +
                    (id_.get("vol") or 0) +
                    (lg.get("vol") or 0)
                )

                # Weighted average price
                if total_mmbtu > 0:
                    w_num = (
                        my_mmbtu  * (my.get("price") or 0) +
                        id_mmbtu  * (id_.get("price") or 0) +
                        lng_mmbtu * (lg.get("price") or 0)
                    )
                    weighted_usd = w_num / total_mmbtu
                    my_share  = 100.0 * my_mmbtu  / total_mmbtu if my_mmbtu  else 0.0
                    id_share  = 100.0 * id_mmbtu  / total_mmbtu if id_mmbtu  else 0.0
                    lng_share = 100.0 * lng_mmbtu / total_mmbtu if lng_mmbtu else 0.0
                else:
                    weighted_usd = None
                    my_share = id_share = lng_share = None

                weighted_sgd = (weighted_usd * fx_usd_sgd) if weighted_usd else None
                implied_floor = (weighted_sgd * 7.5) if weighted_sgd else None

                params = {
                    "price_date":               price_date,
                    "malaysia_vol_mt":           my.get("vol"),
                    "indonesia_vol_mt":          id_.get("vol"),
                    "lng_vol_mt":                lg.get("vol"),
                    "total_vol_mt":              total_vol_mt or None,
                    "malaysia_val_usd":          my.get("val"),
                    "indonesia_val_usd":         id_.get("val"),
                    "lng_val_usd":               lg.get("val"),
                    "malaysia_price_usd_mmbtu":  my.get("price"),
                    "indonesia_price_usd_mmbtu": id_.get("price"),
                    "lng_price_usd_mmbtu":       lg.get("price"),
                    "weighted_avg_usd_mmbtu":    round(weighted_usd, 4) if weighted_usd else None,
                    "malaysia_share_pct":        round(my_share, 2)  if my_share  is not None else None,
                    "indonesia_share_pct":       round(id_share, 2)  if id_share  is not None else None,
                    "lng_share_pct":             round(lng_share, 2) if lng_share is not None else None,
                    "fx_rate_usd_sgd":           fx_usd_sgd,
                    "weighted_avg_sgd_mmbtu":    round(weighted_sgd, 4) if weighted_sgd else None,
                    "implied_usep_floor_sgd_mwh": round(implied_floor, 2) if implied_floor else None,
                    "source":                    "SG Customs / S&P Global",
                    "imported_at":               now_str,
                }

                if dialect == "sqlite":
                    sql = text("""
                        INSERT OR REPLACE INTO gas_prices (
                            price_date,
                            malaysia_vol_mt, indonesia_vol_mt, lng_vol_mt, total_vol_mt,
                            malaysia_val_usd, indonesia_val_usd, lng_val_usd,
                            malaysia_price_usd_mmbtu, indonesia_price_usd_mmbtu, lng_price_usd_mmbtu,
                            weighted_avg_usd_mmbtu,
                            malaysia_share_pct, indonesia_share_pct, lng_share_pct,
                            fx_rate_usd_sgd, weighted_avg_sgd_mmbtu, implied_usep_floor_sgd_mwh,
                            source, imported_at
                        ) VALUES (
                            :price_date,
                            :malaysia_vol_mt, :indonesia_vol_mt, :lng_vol_mt, :total_vol_mt,
                            :malaysia_val_usd, :indonesia_val_usd, :lng_val_usd,
                            :malaysia_price_usd_mmbtu, :indonesia_price_usd_mmbtu, :lng_price_usd_mmbtu,
                            :weighted_avg_usd_mmbtu,
                            :malaysia_share_pct, :indonesia_share_pct, :lng_share_pct,
                            :fx_rate_usd_sgd, :weighted_avg_sgd_mmbtu, :implied_usep_floor_sgd_mwh,
                            :source, :imported_at
                        )
                    """)
                else:
                    sql = text("""
                        INSERT INTO gas_prices (
                            price_date,
                            malaysia_vol_mt, indonesia_vol_mt, lng_vol_mt, total_vol_mt,
                            malaysia_val_usd, indonesia_val_usd, lng_val_usd,
                            malaysia_price_usd_mmbtu, indonesia_price_usd_mmbtu, lng_price_usd_mmbtu,
                            weighted_avg_usd_mmbtu,
                            malaysia_share_pct, indonesia_share_pct, lng_share_pct,
                            fx_rate_usd_sgd, weighted_avg_sgd_mmbtu, implied_usep_floor_sgd_mwh,
                            source, imported_at
                        ) VALUES (
                            :price_date,
                            :malaysia_vol_mt, :indonesia_vol_mt, :lng_vol_mt, :total_vol_mt,
                            :malaysia_val_usd, :indonesia_val_usd, :lng_val_usd,
                            :malaysia_price_usd_mmbtu, :indonesia_price_usd_mmbtu, :lng_price_usd_mmbtu,
                            :weighted_avg_usd_mmbtu,
                            :malaysia_share_pct, :indonesia_share_pct, :lng_share_pct,
                            :fx_rate_usd_sgd, :weighted_avg_sgd_mmbtu, :implied_usep_floor_sgd_mwh,
                            :source, :imported_at
                        )
                        ON CONFLICT (price_date) DO UPDATE SET
                            weighted_avg_usd_mmbtu    = EXCLUDED.weighted_avg_usd_mmbtu,
                            lng_price_usd_mmbtu       = EXCLUDED.lng_price_usd_mmbtu,
                            implied_usep_floor_sgd_mwh = EXCLUDED.implied_usep_floor_sgd_mwh,
                            imported_at               = EXCLUDED.imported_at
                    """)

                conn.execute(sql, params)
                inserted += 1
                all_dates.append(price_date)
            except Exception as e:
                skipped += 1
                if len(errors) < 5:
                    errors.append(f"{year}-{month:02d}: {e}")

    # Summary
    date_range = (min(all_dates), max(all_dates)) if all_dates else (None, None)
    latest_row = None
    try:
        with engine.connect() as conn:
            r = conn.execute(text("""
                SELECT weighted_avg_usd_mmbtu, lng_share_pct, implied_usep_floor_sgd_mwh
                FROM gas_prices
                WHERE weighted_avg_usd_mmbtu IS NOT NULL
                ORDER BY price_date DESC LIMIT 1
            """)).fetchone()
            if r:
                latest_row = r
    except Exception:
        pass

    return {
        "rows_ingested":          inserted,
        "rows_skipped":           skipped,
        "date_range":             date_range,
        "latest_weighted_price":  round(latest_row[0], 2) if latest_row else None,
        "latest_lng_share_pct":   round(latest_row[1], 1) if latest_row else None,
        "latest_implied_usep":    round(latest_row[2], 1) if latest_row else None,
        "errors":                 errors,
    }


def ingest_analyst_forecast(
    file_path,
    engine,
    source_name: str,
    vintage_year: int | None = None,
    granularity: str = "annual",
) -> dict:
    """
    Import external analyst price forecast into forecast_sources + forecast_data.

    granularity options: 'annual' | 'monthly' | 'daily' | 'half_hourly'
    For 'annual': single value replicated to all periods for the year.
    For 'monthly': monthly avg replicated to all periods in each month.
    For 'daily' / 'half_hourly': inserted as-is.

    Expected columns (flexible):
        For annual:      year, price
        For monthly:     year, month, price  OR  date, price
        For daily/hh:    date, period (hh only), price
    """
    file_path = Path(file_path)
    if not file_path.exists():
        return {"rows_imported": 0, "rows_skipped": 0, "errors": [f"File not found: {file_path}"]}

    try:
        if file_path.suffix.lower() in (".xlsx", ".xls"):
            raw = pd.read_excel(file_path, engine="openpyxl")
        else:
            raw = pd.read_csv(file_path)
    except Exception as e:
        return {"rows_imported": 0, "rows_skipped": 0, "errors": [str(e)]}

    # Normalise column names
    col_map = {}
    for c in raw.columns:
        lc = c.lower().strip()
        if lc in ("price", "usep", "forecast_price", "usep_forecast", "annual_avg", "monthly_avg"):
            col_map[c] = "price"
        elif lc in ("year",):
            col_map[c] = "year"
        elif lc in ("month",):
            col_map[c] = "month"
        elif lc in ("date", "price_date"):
            col_map[c] = "date"
        elif lc == "period":
            col_map[c] = "period"
    df = raw.rename(columns=col_map)
    df["price"] = pd.to_numeric(df.get("price"), errors="coerce")
    df = df.dropna(subset=["price"])

    # Expand to per-row records ready for forecast_data
    records: list[dict] = []

    if granularity == "annual":
        for _, row in df.iterrows():
            yr = int(row.get("year", vintage_year or 2025))
            for month in range(1, 13):
                import calendar
                days_in_month = calendar.monthrange(yr, month)[1]
                for day in range(1, days_in_month + 1):
                    dt = f"{yr:04d}-{month:02d}-{day:02d}"
                    records.append({"date": dt, "period": None, "price": float(row["price"])})

    elif granularity == "monthly":
        for _, row in df.iterrows():
            if "date" in df.columns and "date" in row:
                dt = pd.to_datetime(row["date"], errors="coerce")
                yr, mo = dt.year, dt.month
            else:
                yr = int(row.get("year", vintage_year or 2025))
                mo = int(row.get("month", 1))
            import calendar
            days_in_month = calendar.monthrange(yr, mo)[1]
            for day in range(1, days_in_month + 1):
                dt_str = f"{yr:04d}-{mo:02d}-{day:02d}"
                records.append({"date": dt_str, "period": None, "price": float(row["price"])})

    else:  # daily or half_hourly
        for _, row in df.iterrows():
            dt = pd.to_datetime(row.get("date", ""), errors="coerce")
            if pd.isna(dt):
                continue
            period = int(row["period"]) if "period" in df.columns and pd.notna(row.get("period")) else None
            records.append({"date": dt.date().isoformat(), "period": period, "price": float(row["price"])})

    if not records:
        return {"rows_imported": 0, "rows_skipped": 0, "errors": ["No valid rows after expansion"]}

    now_str = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
    dialect = _detect_dialect(engine)

    with engine.begin() as conn:
        result = conn.execute(text("""
            INSERT INTO forecast_sources (source_name, vintage_year, granularity, forecast_type, uploaded_at, row_count)
            VALUES (:name, :vy, :gran, :ftype, :now, :rc)
        """), {
            "name": source_name,
            "vy": vintage_year,
            "gran": granularity,
            "ftype": granularity,
            "now": now_str,
            "rc": len(records),
        })
        source_id = result.lastrowid

    inserted = 0
    skipped = 0
    errors: list[str] = []

    with engine.begin() as conn:
        for rec in records:
            try:
                sql = text("""
                    INSERT INTO forecast_data (source_id, date, period, price)
                    VALUES (:source_id, :date, :period, :price)
                """)
                conn.execute(sql, {"source_id": source_id, "date": rec["date"],
                                   "period": rec["period"], "price": rec["price"]})
                inserted += 1
            except Exception as e:
                skipped += 1
                if len(errors) < 5:
                    errors.append(str(e))

    return {"rows_imported": inserted, "rows_skipped": skipped, "errors": errors, "source_id": source_id}


def ingest_and_retrain(file_path, engine, model_registry_table=None) -> dict:
    """
    Auto-learning pipeline:
      1. Detect file type
      2. Ingest to appropriate table
      3. Check model drift vs last 30 days of forecast_actuals
      4. Retrain XGBoost if drift > 1.5× baseline RMSE
      5. Return summary dict

    Returns: {
        file_type, rows_imported, rows_skipped,
        drift_detected (bool), retrained (bool),
        retrain_metrics (dict or None), errors
    }
    """
    file_path = Path(file_path)
    errors: list[str] = []

    # Step 1: Read & detect
    try:
        if file_path.suffix.lower() in (".xlsx", ".xls"):
            raw = pd.read_excel(file_path, engine="openpyxl")
        else:
            raw = pd.read_csv(file_path)
    except Exception as e:
        return {"file_type": "unknown", "rows_imported": 0, "rows_skipped": 0,
                "drift_detected": False, "retrained": False,
                "retrain_metrics": None, "errors": [str(e)]}

    file_type = detect_file_type(raw)

    # Step 2: Ingest
    if file_type == "nems_prices":
        result = import_csv_to_db(file_path, engine)
    elif file_type == "gas_prices":
        result = ingest_gas_prices(file_path, engine)
    else:
        result = {"rows_imported": 0, "rows_skipped": 0,
                  "errors": [f"Auto-ingest not supported for type: {file_type}"]}
        return {"file_type": file_type, **result,
                "drift_detected": False, "retrained": False, "retrain_metrics": None}

    rows_imported = result.get("rows_imported", 0)
    errors.extend(result.get("errors", []))

    # Steps 3–4: Only attempt for NEMS price files (model relevance)
    drift_detected = False
    retrained = False
    retrain_metrics = None

    if file_type == "nems_prices" and rows_imported > 0:
        try:
            from modules.forecasting import check_model_drift, train_xgboost
            drift_info = check_model_drift(engine, "xgboost_v1", window_days=30)
            drift_detected = drift_info.get("drift_detected", False)

            if drift_detected:
                metrics = train_xgboost(engine)
                retrained = True
                retrain_metrics = metrics
        except Exception as e:
            errors.append(f"Drift/retrain check failed: {e}")

    return {
        "file_type": file_type,
        "rows_imported": rows_imported,
        "rows_skipped": result.get("rows_skipped", 0),
        "drift_detected": drift_detected,
        "retrained": retrained,
        "retrain_metrics": retrain_metrics,
        "errors": errors,
    }


def get_database_summary(engine) -> dict:
    """Return key statistics about the nems_prices table."""
    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT
                MIN(date)        AS min_date,
                MAX(date)        AS max_date,
                COUNT(*)         AS total_rows,
                MAX(imported_at) AS last_import_at
            FROM nems_prices
        """)).mappings().fetchone()

    if row is None or row["total_rows"] == 0:
        return {
            "min_date": None,
            "max_date": None,
            "total_rows": 0,
            "last_import_at": None,
            "coverage_pct": 0.0,
        }

    min_date = row["min_date"]
    max_date = row["max_date"]
    total_rows = row["total_rows"]
    last_import_at = row["last_import_at"]

    from datetime import date as date_type
    if isinstance(min_date, str):
        min_date = date_type.fromisoformat(min_date)
    if isinstance(max_date, str):
        max_date = date_type.fromisoformat(max_date)

    days_in_range = (max_date - min_date).days + 1
    expected_periods = days_in_range * 48
    coverage_pct = round(100.0 * total_rows / expected_periods, 2) if expected_periods > 0 else 0.0

    return {
        "min_date": min_date,
        "max_date": max_date,
        "total_rows": total_rows,
        "last_import_at": last_import_at,
        "coverage_pct": coverage_pct,
    }
